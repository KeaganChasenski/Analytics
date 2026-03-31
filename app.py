import streamlit as st
import openpyxl
import pandas as pd
from collections import defaultdict

# Set up the webpage
st.set_page_config(page_title="Londolozi Ranger Tracker", page_icon="🐆", layout="wide")
st.title("Ranger Daysheet Analyzer")
st.markdown("Upload the monthly Excel daysheet below to automatically calculate total days, private days, and camp splits.")

# 1. Create the file upload button
uploaded_file = st.file_uploader("Upload Monthly Daysheet (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    with st.spinner('Analyzing all 31 tabs & mapping camps... This usually takes 10-15 seconds...'):
        
        # Load the workbook
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        
        # Dictionaries to hold our counts
        total_counts = defaultdict(int)
        private_counts = defaultdict(int)
        camp_counts = defaultdict(lambda: defaultdict(int))
        
        # The 5 main Londolozi camps to look for in Column A
        TARGET_CAMPS = ['TREE', 'VARTY', 'GRANITE', 'FOUNDERS', 'PIONEER']
        
        # Words/characters we want the script to completely ignore
        IGNORE_LIST = ['-', '--', 'nan', 'tbc', 'tba', '?', '??']
        
        # 2. Loop through every day
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rng_cols = set()
            
            # Find the RNG columns
            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and 'RNG' in cell.value.upper():
                        rng_cols.add(cell.column)
            
            daily_rangers = set()
            daily_private = set()
            daily_camp_tracker = defaultdict(set)
            
            # 3. Scan the columns for names, bold text, and Camp sections
            if rng_cols:
                current_camp = None
                
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    
                    # Check the very first column (Column A) to see what camp section we are in
                    col_a_val = row[0].value
                    if col_a_val and isinstance(col_a_val, str):
                        possible_camp = col_a_val.strip().upper()
                        if possible_camp in TARGET_CAMPS:
                            current_camp = possible_camp.title() # Saves as 'Tree', 'Varty', etc.
                    
                    # Now check the RNG columns in this row
                    for cell in row:
                        if cell.column in rng_cols:
                            if cell.value and str(cell.value).upper().strip() != 'RNG':
                                
                                is_bold = cell.font and cell.font.bold
                                
                                for r in str(cell.value).split('/'):
                                    clean_name = r.strip()
                                    
                                    # THE FIX: Check that the name isn't blank, and isn't in our ignore list
                                    if clean_name and clean_name.lower() not in IGNORE_LIST:
                                        
                                        daily_rangers.add(clean_name)
                                        if is_bold:
                                            daily_private.add(clean_name)
                                        
                                        # Log that this ranger drove in this camp today
                                        if current_camp:
                                            daily_camp_tracker[clean_name].add(current_camp)
            
            # 4. Tally up the day across all categories
            for ranger in daily_rangers:
                total_counts[ranger] += 1
            for ranger in daily_private:
                private_counts[ranger] += 1
            for ranger, camps in daily_camp_tracker.items():
                for c in camps:
                    camp_counts[ranger][c] += 1

        # 5. Build the Master Table
        results = []
        all_rangers = set(list(total_counts.keys()) + list(private_counts.keys()))
        
        for ranger in all_rangers:
            results.append({
                "Ranger": ranger,
                "Total Days Driven": total_counts[ranger],
                "Private Days": private_counts[ranger],
                "Tree": camp_counts[ranger].get("Tree", 0),
                "Varty": camp_counts[ranger].get("Varty", 0),
                "Granite": camp_counts[ranger].get("Granite", 0),
                "Founders": camp_counts[ranger].get("Founders", 0),
                "Pioneer": camp_counts[ranger].get("Pioneer", 0)
            })
            
        # Sort by most total days
        df = pd.DataFrame(results).sort_values(by="Total Days Driven", ascending=False).reset_index(drop=True)
        
        st.success("Analysis Complete!")
        
        # Display the table on the website
        st.dataframe(df, use_container_width=True)
        
        # Create a download button for the final numbers
        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download Master Summary (CSV)",
            data=csv,
            file_name='Ranger_Camp_Summary.csv',
            mime='text/csv',
        )
